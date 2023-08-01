#!/usr/bin/evn python
"""
utils for parsing an excel file; includes wrappers for openpyxl
"""
import csv
from os import getcwd, path

from openpyxl import Workbook as wb, load_workbook
from utils.utils import warning, die
from utils.string_utils import xstr
from utils.array_utils import qw

def get_worksheet_names(workbook, print2stderr=False):
    '''
    returns worksheets in a workbook as a list; 
    optional: prints to stderr
    '''
    wsNames = workbook.get_sheet_names()
    if print2stderr:
        warning(str(len(wsNames)), "worksheets found:")
        for name in wsNames:
            warning(name)
    return wsNames


def convert_worksheet_to_csv(worksheet, sep='\t', outputDirectory=None, debug=False):
    '''
    converts specified sheet to csv writes to file with
    same name
    '''
    suffix = '.csv' if sep == ',' else '.txt'
    fileName = worksheet.title.replace(' ', '_') + suffix
    if outputDirectory is None:
        outputDirectory = getcwd()
    fileName = path.join(outputDirectory, fileName)
    lineNum = 0
    with open(fileName, 'w') as f:
        writer = csv.writer(f, delimiter=sep)
        for row in worksheet.rows:
            if debug:
                lineNum = lineNum + 1
                warning(lineNum, ":", [cell.value for cell in row])
            writer.writerow([cell.value for cell in row])


def get_column_names(worksheet, rownum):
    '''
    extract column names from specified row
    '''
    fields = [cell.value for cell in tuple(worksheet.rows)[rownum - 1]]
    return fields


def extract_field_values(worksheet, field, headerRow=1):
    '''
    extract field (column values) from a worksheet
    '''
    columns = get_column_names(worksheet, headerRow)
    index = columns.index(field)
    values = []
    for row in worksheet.iter_rows(min_row=headerRow + 1):
        values.append(row[index].value)
                                  
    return values


def get_worksheet_from_file(fileName, wsName):
    '''load worksheet from an excel file'''
    workbook = load_workbook(fileName, data_only=True)
    return workbook.get_sheet_by_name(name=wsName)

    
def load_workbook_from_file(fileName):
    '''load workbook from an excel file'''
    return load_workbook(fileName, data_only=True)
