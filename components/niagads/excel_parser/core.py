import logging
import json

from os import path
from typing import Union
from pandas import read_excel, DataFrame
from openpyxl import Workbook as wb, load_workbook

from niagads.dict_utils.core import convert_str2numeric_values
from niagads.string_utils.core import xstr, to_snake_case
from niagads.pd_dataframe.core import strip

class ExcelFileParser:
    """
    parser for EXCEL files; mainly a pandas/openpyxl wrapper
    """
    def __init__(self, file:str, debug:bool=False):
        """
        init new ExcelFileParser

        Args:
            file (str): EXCEL file name (full path)
            debug (bool, optional): enable debug mode. Defaults to False.
        """
        self._debug = debug
        self.logger = logging.getLogger(__name__)
        self.__file = file
        self.__na = None # missing value string representation
        self.__strip = False # flag for trimming leading & trailing whitespace
        
        # openpyxl data structures
        # useful for iterating over sheets
        self.__workbook = None
        self.__worksheets = None
        self.__init_workbook()
        
        
    def __init_workbook(self):
        """
        load the workbook & get list of worksheets
        initialize member variables
        (should do basic parsing & error reporting; e.g., non-ascii characters)   
        """
        # data_only=True -> values, not formulas & formatting
        self.__workbook = load_workbook(self.__file, data_only=True)
        self.__worksheets = self.__workbook.sheetnames
 
 
    def na(self, value: str):
        """
        fill NA's with specified value when using pandas conversions

        Args:
            value (str): value to fill (e.g., 'NULL', 'NA', '.')
        """
        self.__na = value
 
 
    def strip(self, strip=True):
        """
        flag indicating whether to iterate over all fields and 
        trim leading and trailing spaces when converting to JSON or CSV

        Args:
            strip (bool, optional): trim leading and trailing spaces from all fields. Defaults to True.
        """
        self.__strip = strip
        

    def set_na_rep(self, value:str):
        """
        set the NA (missing / null) value string representation
        suggestions: '.', 'NA', 'null', 'NULL', '\t'
        
        currently only applies to CSV export

        Args:
            value (str): missing / null value string
        """
        self.__na = value
        
        
    def get_workbook(self):
        """
        returns openpyxl `workbook` object
        """
        return self.__workbook
    
    
    def list_worksheets(self):
        """
        returns a list of worksheet names
        """
        return self.__worksheets
    
    
    def write_to_csv(self, outputPath:str, worksheet:str=None, sep:str=','):
        """
        converts the EXCEL file to CSV and saved to `outputPath`.  
        If `worksheet` == None, all worksheets in the file are converted
        otherwise only the specified worksheet will be converted

        output files will be named `to_snake_case(worksheet)` +'.csv'` 
        '.txt' extension will be used for space and tab delimited files
        '.csv' will be use for comma and non-standard delimiters     

        Args:
            outputPath (str): path to which the files should be written
            sep (str, optional): delimiter.  Defaults to ','
            worksheet (str, optional): worksheet name. Defaults to None.
        """
        if worksheet is None:
            for ws in self.__worksheets: self.__worksheet_to_csv(outputPath, ws, sep)
        else:
            self.__worksheet_to_csv(outputPath, worksheet, sep)
        
            
            
    def is_valid_worksheet(self, worksheet:str, raiseErr:bool=False):
        """
        checks if worksheet exists, otherwise raises error

        Args:
            worksheet (str): worksheet name
            raiseErr (boolean, optional): raise an error if False, otherwise will return False
        """
        if worksheet not in self.__worksheets:
            if raiseErr:
                msg = "Worksheet " + worksheet + " not found in " \
                    + self.__file + "; valid values are: " + xstr(self.__worksheets)
                self.logger.error(msg)
                raise ValueError(msg)
            else:
                return False
        return True
    
    
    def __worksheet_to_csv(self, outputPath:str, worksheet:str, sep:str):
        """
        internal / called by to_csv to write an worksheet to CSV

        Args:
            outputPath (str): path to which the files should be written
            sep (str): delimiter
            worksheet (str): worksheet name
        """        
        fileName = path.join(outputPath, 
                                to_snake_case(worksheet) \
                                + ('.csv' if sep == ',' else '.txt'))
        
        df = self.to_pandas_df(worksheet)
        df.to_csv(fileName, sep=sep, index=False, encoding='utf-8', na_rep=self.__na)
    
    
    def worksheet_to_json(self, worksheet:Union[str, int], transpose=False, returnStr=False, **kwargs):
        """
        converts the EXCEL file to JSON
        
        Args:
            worksheet (str|int): worksheet name or index, starting from 0.
            transpose (bool, optional): transpose the worksheet?
            returnStr (bool, optional): return jsonStr instead of object
            **kwargs (optional): arguments to pass to `pandas` `read_excel` see
                (see https://pandas.pydata.org/docs/reference/api/pandas.read_excel.html))
        
        Returns:
            if `returnStr` returns JSON string instead of object
        """
        ws = self.__worksheets[worksheet] if isinstance(worksheet, int) else worksheet   
        
        # orient='records' returns indexes; e.g. [index: {row data}] so need to extract the values
        jsonStr = self.to_pandas_df(ws, transpose, **kwargs).to_json(orient = 'records')
        
        # convert strings to numeric so can do typing validation
        jsonObj = json.loads(jsonStr)
        if isinstance(jsonObj, list):
            jsonObj = [convert_str2numeric_values(r) for r in json.loads(jsonStr)]
        else:
            jsonObj = convert_str2numeric_values(jsonObj)
        
        return json.dumps(jsonObj) if returnStr else json.loads(jsonStr)


    def __trim(self, df: DataFrame):
        """
        trims trailing spaces if set in options

        Args:
            df (DataFrame): pandas data frame
        """
        return strip(df) if self.__strip else df
    
        
    def to_pandas_df(self, worksheet: str, transpose=False, **kwargs) -> DataFrame:
        """
        _summary_

        Args:
            worksheet (str): worksheet name
            transpose (str): transpose the worksheet
            **kwargs: must match expected args for pandas.read_excel 
                (see https://pandas.pydata.org/docs/reference/api/pandas.read_excel.html)

        Returns:
            DataFrame: worksheet in data frame format
        """
        # raise error if False
        self.is_valid_worksheet(worksheet, raiseErr=True)
        df: DataFrame = read_excel(self.__file, sheet_name=worksheet, **kwargs)
        if self.__na is not None:
            df.fillna(self.__na)
        return self.__trim(df.T) if transpose else self.__trim(df)
